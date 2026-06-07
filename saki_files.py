import os
import datetime
import tempfile
import shutil
import logging
from typing import Optional, Tuple

import ollama

from saki_database import simpan_dokumen, tambah_ke_chroma, DOCUMENTS_FOLDER

logger = logging.getLogger("saki")

# Local config
SUMMARY_MAX_LENGTH = int(os.getenv("SUMMARY_MAX_LENGTH", 8000))
MODEL = os.getenv("MODEL", "qwen3:4b")
SYSTEM_PROMPT = os.getenv("SYSTEM_PROMPT", "") or "Ringkas dokumen berikut." 


def ekstrak_teks_dari_pdf(filepath: str) -> Optional[str]:
    from PyPDF2 import PdfReader
    try:
        reader = PdfReader(filepath)
        text = "\n".join([page.extract_text() or "" for page in reader.pages])
        if not text.strip():
            logger.warning(f"PDF kosong: {filepath}")
            return None
        logger.info(f"Berhasil ekstrak PDF: {os.path.basename(filepath)} ({len(text)} chars)")
        return text.strip()
    except FileNotFoundError:
        logger.error(f"PDF tidak ditemukan: {filepath}")
        return None
    except Exception as e:
        logger.error(f"Gagal ekstrak PDF {os.path.basename(filepath)}: {type(e).__name__}: {str(e)}", exc_info=True)
        return None


def ekstrak_teks_dari_docx(filepath: str) -> Optional[str]:
    from docx import Document
    try:
        doc = Document(filepath)
        text = "\n".join([para.text for para in doc.paragraphs])
        if not text.strip():
            logger.warning(f"DOCX kosong: {filepath}")
            return None
        logger.info(f"Berhasil ekstrak DOCX: {os.path.basename(filepath)} ({len(text)} chars)")
        return text.strip()
    except FileNotFoundError:
        logger.error(f"DOCX tidak ditemukan: {filepath}")
        return None
    except Exception as e:
        logger.error(f"Gagal ekstrak DOCX {os.path.basename(filepath)}: {type(e).__name__}: {str(e)}", exc_info=True)
        return None


def ekstrak_teks_dari_txt(filepath: str) -> Optional[str]:
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            text = f.read()
        if not text.strip():
            logger.warning(f"File kosong: {filepath}")
            return None
        logger.info(f"Berhasil baca file: {os.path.basename(filepath)} ({len(text)} chars)")
        return text.strip()
    except UnicodeDecodeError:
        try:
            with open(filepath, "r", encoding="latin-1") as f:
                text = f.read()
            logger.warning(f"File dibaca dengan encoding latin-1: {filepath}")
            return text.strip()
        except Exception as e:
            logger.error(f"Gagal baca file dengan encoding alternatif: {str(e)}", exc_info=True)
            return None
    except FileNotFoundError:
        logger.error(f"File tidak ditemukan: {filepath}")
        return None
    except Exception as e:
        logger.error(f"Gagal baca file {os.path.basename(filepath)}: {type(e).__name__}: {str(e)}", exc_info=True)
        return None

def ekstrak_teks_dari_csv(filepath: str) -> Optional[str]:
    """Ekstrak teks dari file CSV."""
    import csv
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)
        
        if not rows:
            logger.warning(f"CSV kosong: {filepath}")
            return None
        
        text = ""
        for row in rows[:501]:
            text += " | ".join(row) + "\n"
        
        logger.info(f"Berhasil eks`trak CSV: {os.path.basename(filepath)} ({len(rows)} baris)")
        return text.strip()
    except UnicodeDecodeError:
        try:
            with open(filepath, "r", encoding="latin-1") as f:
                reader = csv.reader(f)
                rows = list(reader)
            text = "\n".join([" | ".join(row) for row in rows[:501]])
            return text.strip()
        except Exception as e:
            logger.error(f"Gagal ekstrak CSV: {str(e)}")
            return None
    except Exception as e:
        logger.error(f"Gagal ekstrak CSV {os.path.basename(filepath)}: {type(e).__name__}: {str(e)}", exc_info=True)
        return None


def ekstrak_teks_dari_xlsx(filepath: str) -> Optional[str]:
    """Ekstrak teks dari file Excel (.xlsx)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        
        all_text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_text = [f"=== Sheet: {sheet_name} ==="]
            
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                if row_count > 100:
                    sheet_text.append("... (dipotong)")
                    break
                sheet_text.append(" | ".join([str(cell) if cell is not None else "" for cell in row]))
                row_count += 1
            
            all_text.append("\n".join(sheet_text))
        
        wb.close()
        text = "\n\n".join(all_text)
        
        if not text.strip():
            logger.warning(f"Excel kosong: {filepath}")
            return None
        
        logger.info(f"Berhasil ekstrak Excel: {os.path.basename(filepath)}")
        return text.strip()
    
    except ImportError:
        logger.error("openpyxl tidak terinstall. Install dengan: pip install openpyxl")
        return None
    except Exception as e:
        logger.error(f"Gagal ekstrak Excel {os.path.basename(filepath)}: {type(e).__name__}: {str(e)}", exc_info=True)
        return None


def proses_upload(uploaded_file) -> Tuple[Optional[int], str]:
    """Proses upload file: ekstrak, ringkas via Ollama, simpan ke DB dan Chroma."""
    tmp_path = None
    try:
        filename = uploaded_file.name
        ext = os.path.splitext(filename)[1].lower()

        if ext not in [".pdf", ".docx", ".txt", ".md", ".csv", ".xlsx", ".xls"]:
            return None, f"Format {ext} tidak didukung. Gunakan PDF, DOCX, TXT, MD, CSV, atau Excel."

        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            tmp.write(uploaded_file.read())
            tmp_path = tmp.name

        logger.info(f"Processing upload: {filename}")

        if ext == ".pdf":
            teks = ekstrak_teks_dari_pdf(tmp_path)
            file_type = "PDF"
        elif ext == ".docx":
            teks = ekstrak_teks_dari_docx(tmp_path)
            file_type = "DOCX"
        elif ext in [".txt", ".md"]:
            teks = ekstrak_teks_dari_txt(tmp_path)
            file_type = "TXT"
        elif ext == ".csv":
            teks = ekstrak_teks_dari_csv(tmp_path)
            file_type = "CSV"
        elif ext in [".xlsx", ".xls"]:
            teks = ekstrak_teks_dari_xlsx(tmp_path)
            file_type = "Excel"
        else:
            return None, f"Format {ext} tidak didukung."

        if not teks:
            logger.warning(f"Ekstraksi teks kosong: {filename}")
            return None, "Gagal mengekstrak teks. File mungkin kosong atau corrupt."

        # V8.1: Untuk CSV besar, buat ringkasan statistik
        if ext in [".csv", ".xlsx", ".xls"]:
            lines = teks.split('\n')
            jumlah_baris = len(lines)
            header = lines[0] if lines else ""
            
            # Info statistik
            teks_ringkas = f"File CSV: {filename}\n"
            teks_ringkas += f"Jumlah total baris: {jumlah_baris}\n"
            teks_ringkas += f"Header: {header}\n\n"
            teks_ringkas += f"10 baris pertama:\n"
            teks_ringkas += '\n'.join(lines[:11])  # header + 10 baris
            teks_ringkas += f"\n\n10 baris terakhir:\n"
            teks_ringkas += '\n'.join(lines[-10:])
            teks_ringkas = teks_ringkas[:SUMMARY_MAX_LENGTH]
        else:
            teks_ringkas = teks[:SUMMARY_MAX_LENGTH]       

        # Prompt dibuat di luar if-else (berlaku untuk semua format)
        prompt = f"Ringkas dokumen berikut:\n\nJudul: {filename}\n\n{teks_ringkas}"
        try:
            response = ollama.chat(model=MODEL, messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": prompt}
            ])
            ringkasan = response["message"]["content"]
        except Exception as e:
            logger.error(f"Ollama summarization failed: {str(e)}", exc_info=True)
            ringkasan = teks_ringkas[:1000]

        os.makedirs(DOCUMENTS_FOLDER, exist_ok=True)
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        saved_filename = f"{timestamp}_{filename}"
        saved_path = os.path.join(DOCUMENTS_FOLDER, saved_filename)

        shutil.copy2(tmp_path, saved_path)

        doc_id = simpan_dokumen(filename, saved_path, file_type, teks, ringkasan)
        if doc_id:
            try:
                tambah_ke_chroma(doc_id, teks, filename)
            except Exception as e:
                logger.error(f"Failed to add to Chroma: {str(e)}", exc_info=True)

        logger.info(f"Upload berhasil: {filename} -> doc_id={doc_id}")
        return doc_id, ringkasan
    except Exception as e:
        logger.error(f"Error processing upload: {type(e).__name__}: {str(e)}", exc_info=True)
        return None, f"Gagal memproses file: {type(e).__name__}"
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
                logger.debug(f"Temp file cleaned: {tmp_path}")
            except Exception as e:
                logger.warning(f"Failed to cleanup temp file {tmp_path}: {str(e)}")
